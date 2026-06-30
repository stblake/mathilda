"""Tests for generate.py — build, save, re-read, assert."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from generate import build_workbook_from_spec, write_workbook


def _sample_spec() -> dict:
    return {
        "sheets": [
            {
                "name": "Summary",
                "title": "Q3 summary",
                "columns": ["Metric", "Actual", "Plan"],
                "rows": [
                    ["Revenue", 124, 118],
                    ["Costs", 56, 58],
                ],
                "number_formats": {"Actual": "#,##0", "Plan": "#,##0"},
            },
            {
                "name": "Detail",
                "columns": ["Segment", "ARR"],
                "rows": [["Enterprise", 8400], ["SMB", 2100]],
            },
        ]
    }


def test_builds_two_sheets(tmp_path: Path) -> None:
    wb = build_workbook_from_spec(_sample_spec())
    out = tmp_path / "book.xlsx"
    wb.save(out)
    reread = load_workbook(out)
    assert reread.sheetnames == ["Summary", "Detail"]


def test_title_row_written(tmp_path: Path) -> None:
    wb = build_workbook_from_spec(_sample_spec())
    out = tmp_path / "book.xlsx"
    wb.save(out)
    reread = load_workbook(out)
    ws = reread["Summary"]
    assert ws.cell(row=1, column=1).value == "Q3 summary"
    assert ws.cell(row=3, column=1).value == "Metric"   # header row after title


def test_data_rows_written(tmp_path: Path) -> None:
    wb = build_workbook_from_spec(_sample_spec())
    out = tmp_path / "book.xlsx"
    wb.save(out)
    reread = load_workbook(out)
    ws = reread["Summary"]
    assert ws.cell(row=4, column=1).value == "Revenue"
    assert ws.cell(row=4, column=2).value == 124


def test_number_format_applied(tmp_path: Path) -> None:
    wb = build_workbook_from_spec(_sample_spec())
    out = tmp_path / "book.xlsx"
    wb.save(out)
    reread = load_workbook(out)
    ws = reread["Summary"]
    assert ws.cell(row=4, column=2).number_format == "#,##0"


def test_freeze_panes_set(tmp_path: Path) -> None:
    wb = build_workbook_from_spec(_sample_spec())
    out = tmp_path / "book.xlsx"
    wb.save(out)
    reread = load_workbook(out)
    ws = reread["Summary"]
    assert ws.freeze_panes is not None


def test_empty_spec_rejected() -> None:
    with pytest.raises(ValueError):
        build_workbook_from_spec({})


def test_write_workbook_from_dataframes(tmp_path: Path) -> None:
    pd = pytest.importorskip("pandas")
    df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
    out = tmp_path / "df.xlsx"
    write_workbook({"Numbers": df}, out, title_row=False)
    reread = load_workbook(out)
    ws = reread["Numbers"]
    assert ws.cell(row=1, column=1).value == "a"
    assert ws.cell(row=2, column=2).value == 3


def test_sheet_name_sanitized() -> None:
    wb = build_workbook_from_spec({
        "sheets": [
            {"name": "bad/chars?*", "columns": ["a"], "rows": [[1]]},
        ]
    })
    # Should not raise; chars replaced.
    assert wb.sheetnames[0] == "bad_chars__"
