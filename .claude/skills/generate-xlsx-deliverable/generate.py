"""Generate a styled multi-sheet .xlsx deliverable.

Usage (CLI):
    python generate.py <spec.json|spec.yaml> [--output out.xlsx]

Usage (Python):
    from generate import write_workbook
    write_workbook({"Sheet1": df}, "out.xlsx", title_row=True)

Spec schema (JSON or YAML):
    sheets:
      - name: Summary              # sheet tab name
        title: Optional title row
        columns: [col1, col2, ...]
        rows:
          - [v1, v2, ...]
        number_formats:            # optional
          col2: "#,##0"
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F3B5C")   # default navy header
HEADER_FONT = Font(bold=True, color="FFFFFF")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F4F7")
TITLE_FONT = Font(bold=True, size=14, color="1F3B5C")
MAX_COL_WIDTH = 48


def load_spec(path: Path) -> dict[str, Any]:
    text = path.read_text()
    if path.suffix.lower() in {".yaml", ".yml"}:
        import yaml
        return yaml.safe_load(text)
    return json.loads(text)


def build_workbook_from_spec(spec: dict[str, Any]) -> Workbook:
    wb = Workbook()
    # Remove the default sheet; we'll add ours.
    wb.remove(wb.active)

    sheets = spec.get("sheets") or []
    if not sheets:
        raise ValueError("spec must contain at least one sheet")

    for s in sheets:
        _write_sheet(
            wb,
            name=s.get("name") or "Sheet",
            title=s.get("title"),
            columns=s.get("columns") or [],
            rows=s.get("rows") or [],
            number_formats=s.get("number_formats") or {},
        )
    return wb


def write_workbook(sheets, output: str | Path, *, title_row: bool = False) -> Path:
    """Write a dict[name, DataFrame] to a styled workbook.

    If title_row is True, each sheet tab name is also used as an H1 title row.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for name, df in sheets.items():
        _write_sheet(
            wb,
            name=name,
            title=name if title_row else None,
            columns=list(df.columns),
            rows=df.values.tolist(),
            number_formats={},
        )
    out = Path(output)
    wb.save(out)
    return out


def _write_sheet(wb, *, name, title, columns, rows, number_formats):
    ws = wb.create_sheet(title=_safe_sheet_name(name))
    start_row = 1

    if title:
        ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
        if len(columns) > 1:
            ws.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=start_row,
                end_column=max(len(columns), 1),
            )
        start_row += 2  # title + blank line

    header_row = start_row
    for c, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r, row in enumerate(rows, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if (r - header_row) % 2 == 0:
                cell.fill = ZEBRA_FILL
        # Apply per-column number format.
        for c, col in enumerate(columns, start=1):
            fmt = number_formats.get(col)
            if fmt:
                ws.cell(row=r, column=c).number_format = fmt

    _autosize_columns(ws, columns, rows, header_row)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _autosize_columns(ws, columns, rows, header_row) -> None:
    widths: dict[int, int] = {}
    for c, col in enumerate(columns, start=1):
        widths[c] = min(len(str(col)) + 2, MAX_COL_WIDTH)
    for row in rows:
        for c, val in enumerate(row, start=1):
            widths[c] = max(widths.get(c, 0), min(len(str(val)) + 2, MAX_COL_WIDTH))
    for c, width in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = width


def _safe_sheet_name(name: str) -> str:
    # Excel rejects these chars in sheet names, and caps at 31 chars.
    for bad in r'\/?*[]:':
        name = name.replace(bad, "_")
    return name[:31] or "Sheet"


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Generate a styled .xlsx from a spec.")
    ap.add_argument("spec", help="Path to a JSON or YAML spec.")
    ap.add_argument("--output", "-o", default="deliverable.xlsx")
    args = ap.parse_args(argv)

    spec_path = Path(args.spec)
    if not spec_path.exists():
        print(f"error: spec not found: {spec_path}", file=sys.stderr)
        return 2

    wb = build_workbook_from_spec(load_spec(spec_path))
    out = Path(args.output)
    wb.save(out)
    print(f"wrote {out} ({len(wb.sheetnames)} sheet(s))")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
